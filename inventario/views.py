from django.utils import timezone  
from datetime import timedelta 
from django.forms import ValidationError
from django.shortcuts import get_object_or_404, render, redirect
from django.views.generic.edit import CreateView
from django.urls import reverse_lazy
from authentication import models
from inventario.services import descontar_lotes_para_pieza
from .models import HistorialPrecio, Pieza, MovimientoInventario, Lote, Categoria, Proveedor, Ubicacion, Kit, KitItem
from .forms import KitForm, KitItemFormSet, MovimientoInventarioForm,LoteForm, HistorialPrecioForm, PiezaForm, UbicacionForm,CategoriaForm, ProveedorForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from authentication.decorators import group_required  # ajusta el path según tu estructura
from django.db.models import F 
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.template.loader import get_template
from xhtml2pdf import pisa

# Create your views here.
@login_required   
def dashboard_view(request):
    group = request.user.groups.first()
    role_display = group.name if group else "Sin rol"

    total_piezas = Pieza.objects.count()
    piezas_stock_bajo = Pieza.objects.filter(cantidad__lte=F("stock_minimo")).count()
    movimientos_totales = MovimientoInventario.objects.count()
    ubicaciones_distintas = Pieza.objects.values('ubicacion').distinct().count()

    context = {
        'user': request.user,
        'role_display': role_display,
        'total_piezas': total_piezas,
        'piezas_stock_bajo': piezas_stock_bajo,
        'movimientos_totales': movimientos_totales,
        'ubicaciones_distintas': ubicaciones_distintas
    }

    return render(request, 'dashboards/dashboard.html', context)

def home(request):
    return render(request, 'inventario/home.html')

class PiezaCreateView(CreateView):
    model = Pieza
    form_class = PiezaForm
    template_name = 'inventario/registrar_pieza.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        self.object = form.save()
        return super().form_valid(form)


def lista_piezas(request): # Cambiamos el nombre para reflejar 'Pieza'
    piezas_list = Pieza.objects.all() # Usamos el modelo Pieza
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.all()

    # Obtener los parámetros de filtrado de la URL (GET request)
    categoria_id = request.GET.get('categoria')
    ubicacion_id = request.GET.get('ubicacion')

    # Aplicar filtros si los parámetros están presentes
    if categoria_id:
        piezas_list = piezas_list.filter(categoria__id=categoria_id) # Filtra por el ID de la categoría de la Pieza

    if ubicacion_id:
        piezas_list = piezas_list.filter(ubicacion__id=ubicacion_id) # Filtra por el ID de la ubicación de la Pieza

    context = {
        'piezas': piezas_list, # Cambiamos 'productos' por 'piezas'
        'categorias': categorias,
        'ubicaciones': ubicaciones,
        'categoria_seleccionada': categoria_id,
        'ubicacion_seleccionada': ubicacion_id
    }
    return render(request, 'inventario/listar_piezas.html', context) # Ajusta el path a tu template

def exportar_reporte_excel(request):
    hoy = timezone.now().date()

    piezas_stock_bajo = Pieza.objects.filter(cantidad__lte=models.F('stock_minimo'))
    lotes_vencidos = Lote.objects.filter(fecha_vencimiento__lt=hoy, pieza__requiere_vencimiento=True)
    lotes_por_vencer = Lote.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timezone.timedelta(days=7),
        pieza__requiere_vencimiento=True
    )
    movimientos = MovimientoInventario.objects.all().order_by('-fecha')[:20]

    # Crear workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Stock Bajo"

    # Stock bajo
    ws1.append(["Pieza", "Cantidad", "Stock Mínimo"])
    for pieza in piezas_stock_bajo:
        ws1.append([pieza.nombre, pieza.cantidad, pieza.stock_minimo])

    # Lotes vencidos
    ws2 = wb.create_sheet(title="Lotes Vencidos")
    ws2.append(["Pieza", "Código Lote", "Fecha Vencimiento"])
    for lote in lotes_vencidos:
        ws2.append([lote.pieza.nombre, lote.codigo_lote, lote.fecha_vencimiento])

    # Lotes por vencer
    ws3 = wb.create_sheet(title="Lotes por Vencer")
    ws3.append(["Pieza", "Código Lote", "Fecha Vencimiento"])
    for lote in lotes_por_vencer:
        ws3.append([lote.pieza.nombre, lote.codigo_lote, lote.fecha_vencimiento])

    # Movimientos
    ws4 = wb.create_sheet(title="Últimos Movimientos")
    ws4.append(["Fecha", "Tipo", "Pieza", "Cantidad", "Responsable"])
    for mov in movimientos:
        ws4.append([
            mov.fecha.strftime("%Y-%m-%d %H:%M"),
            mov.get_tipo_display(),
            mov.pieza.nombre,
            mov.cantidad,
            mov.responsable.get_full_name() if mov.responsable else "-"
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
    wb.save(response)
    return response

@login_required
def listar_ubicaciones(request):
    ubicaciones = Ubicacion.objects.all()
    return render(request, 'ubicaciones/listar.html', {'ubicaciones': ubicaciones})

@login_required
def crear_ubicacion(request):
    form = UbicacionForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('listar_ubicaciones')
    return render(request, 'ubicaciones/form.html', {'form': form, 'accion': 'Crear'})

@login_required
def editar_ubicacion(request, pk):
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    form = UbicacionForm(request.POST or None, instance=ubicacion)
    if form.is_valid():
        form.save()
        return redirect('listar_ubicaciones')
    return render(request, 'inventario/form.html', {'form': form, 'accion': 'Editar'})


@login_required
def listar_categorias(request):
    categorias = Categoria.objects.all()
    return render(request, 'categorias/listar.html', {'categorias': categorias})

@login_required
def crear_categoria(request):
    form = CategoriaForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('listar_categorias')
    return render(request, 'categorias/form.html', {'form': form, 'accion': 'Crear'})

@login_required
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    form = CategoriaForm(request.POST or None, instance=categoria)
    if form.is_valid():
        form.save()
        return redirect('listar_categorias')
    return render(request, 'categorias/form.html', {'form': form, 'accion': 'Editar'})



@login_required
def eliminar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)

    # Verifica si hay piezas asociadas
    piezas_asociadas = categoria.pieza_set.count()  # type: ignore # relaciona con el modelo Pieza

    if request.method == 'POST':
        if piezas_asociadas > 0:
            messages.error(request, 'No se puede eliminar la categoría porque hay piezas asociadas.')
            return redirect('listar_categorias')
        categoria.delete()
        messages.success(request, 'Categoría eliminada correctamente.')
        return redirect('listar_categorias')

    return render(request, 'categorias/eliminar_confirmacion.html', {
        'categoria': categoria,
        'piezas_asociadas': piezas_asociadas
    })


def pieza_detalle(request, pk):
    pieza = get_object_or_404(Pieza, pk=pk)
    return render(request, 'inventario/detalle_piezas.html', {'pieza': pieza})

@login_required
def registrar_movimiento(request):
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)

            if movimiento.tipo == 'salida':
                movimiento.save()  # guarda primero para tener el ID
                try:
                    descontar_lotes_para_pieza(
                        pieza=movimiento.pieza,
                        cantidad_requerida=movimiento.cantidad,
                        movimiento=movimiento
                    )
                except ValidationError as e:
                    form.add_error(None, str(e))
                    return render(request, 'movimientos/form.html', {'form': form})

            else:
                movimiento.save()

            messages.success(request, "Movimiento registrado correctamente.")
            return redirect('listar_movimientos')
    else:
        form = MovimientoInventarioForm()

    return render(request, 'movimientos/form.html', {'form': form})

def registrar_lote(request):
    if request.method == 'POST':
        form = LoteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Lote agregado correctamente.")
            return redirect('listar_lotes')
    else:
        form = LoteForm()
    return render(request, 'lotes/form.html', {'form': form, 'titulo': 'Registrar nuevo lote'})

@login_required
def listar_lotes(request):
    lotes = Lote.objects.select_related('pieza').order_by('-fecha_emision')
    return render(request, 'lotes/listar.html', {'lotes': lotes})


def registrar_precio(request):
    pieza_id = request.GET.get('pieza_id')
    pieza = Pieza.objects.filter(pk=pieza_id).first()

    if request.method == 'POST':
        form = HistorialPrecioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('pieza_detalle', pk=form.cleaned_data['pieza'].id)
    else:
        form = HistorialPrecioForm(initial={'pieza': pieza}) if pieza else HistorialPrecioForm()

    return render(request, 'registrar_precio.html', {'form': form})

def editar_pieza(request, pk):
    pieza = get_object_or_404(Pieza, pk=pk)
    if request.method == 'POST':
        form = PiezaForm(request.POST, request.FILES, instance=pieza)
        if form.is_valid():
            form.save()
            return redirect('pieza_detalle', pk=pieza.pk)
    else:
        form = PiezaForm(instance=pieza)
    return render(request, 'inventario/editar_pieza.html', {'form': form, 'pieza': pieza})

@group_required(['Administrador del Sistema', 'Gestor de Inventario'])
def registrar_pieza(request):
    if request.method == 'POST':
        form = PiezaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pieza registrada correctamente.')
            return redirect('listar_piezas')  # Asegúrate de tener esta vista o cámbiala
    else:
        form = PiezaForm()
    return render(request, 'inventario/registrar_pieza.html', {'form': form})


@login_required
def listar_proveedores(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'proveedores/listar.html', {'proveedores': proveedores})


@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'proveedores/form.html', {'form': form, 'titulo': 'Agregar Proveedor'})

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('listar_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'proveedores/form.html', {'form': form, 'titulo': 'Editar Proveedor'})


@login_required
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('listar_proveedores')
    return render(request, 'proveedores/confirmar_eliminar.html', {'proveedor': proveedor})

def listar_kits(request):
    kits = Kit.objects.all()
    return render(request, 'kits/listar.html', {'kits': kits})

@login_required
def crear_kit(request):
    if request.method == 'POST':
        form = KitForm(request.POST)
        formset = KitItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            kit = form.save()
            formset.instance = kit
            formset.save()
            return redirect('listar_kits')
    else:
        form = KitForm()
        formset = KitItemFormSet()
    return render(request, 'kits/form.html', {'form': form, 'formset': formset, 'titulo': 'Crear Kit'})

@login_required
def editar_kit(request, pk):
    kit = get_object_or_404(Kit, pk=pk)
    if request.method == 'POST':
        form = KitForm(request.POST, instance=kit)
        formset = KitItemFormSet(request.POST, instance=kit)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('listar_kits')
    else:
        form = KitForm(instance=kit)
        formset = KitItemFormSet(instance=kit)
    return render(request, 'kits/form.html', {'form': form, 'formset': formset, 'titulo': 'Editar Kit'})

@login_required
def eliminar_kit(request, pk):
    kit = get_object_or_404(Kit, pk=pk)
    if request.method == 'POST':
        kit.delete()
        return redirect('listar_kits')
    return render(request, 'kits/confirmar_eliminar.html', {'kit': kit})


@login_required
def registrar_salida_kit(request):
    if request.method == 'POST':
        kit_id = request.POST.get('kit')
        cantidad_kits = int(request.POST.get('cantidad', 1))
        kit = Kit.objects.get(id=kit_id)

        try:
            for item in kit.kititem_set.all():
                total_cantidad = item.cantidad * cantidad_kits
                pieza = item.pieza

                # Crear movimiento por cada pieza
                movimiento = MovimientoInventario.objects.create(
                    pieza=pieza,
                    tipo='salida',
                    cantidad=total_cantidad,
                    responsable=request.user,
                    observacion=f'Salida a través del kit: {kit.nombre}'
                )

                descontar_lotes_para_pieza(
                    pieza=pieza,
                    cantidad_requerida=total_cantidad,
                    movimiento=movimiento
                )

            messages.success(request, f"Salida del kit '{kit.nombre}' registrada correctamente.")
            return redirect('listar_kits')

        except ValidationError as e:
            messages.error(request, str(e))

    kits = Kit.objects.all()
    return render(request, 'kits/registrar_salida.html', {'kits': kits})
def reportes_personalizados(request):
    hoy = timezone.now().date() # type: ignore

    piezas_stock_bajo = Pieza.objects.filter(cantidad__lte=models.F('stock_minimo'))

    lotes_vencidos = Lote.objects.filter(
        fecha_vencimiento__lt=hoy,
        pieza__requiere_vencimiento=True
    )

    lotes_por_vencer = Lote.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timezone.timedelta(days=7), # type: ignore
        pieza__requiere_vencimiento=True
    )

    movimientos = MovimientoInventario.objects.all().order_by('-fecha')[:20]

    return render(request, 'reportes/index.html', {
        'piezas_stock_bajo': piezas_stock_bajo,
        'lotes_vencidos': lotes_vencidos,
        'lotes_por_vencer': lotes_por_vencer,
        'movimientos': movimientos,
    })


@login_required
def exportar_reporte_pdf(request):
    hoy = timezone.now().date()

    piezas_stock_bajo = Pieza.objects.filter(cantidad__lte=F('stock_minimo'))
    lotes_vencidos = Lote.objects.filter(fecha_vencimiento__lt=hoy, pieza__requiere_vencimiento=True)
    lotes_por_vencer = Lote.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timezone.timedelta(days=7),
        pieza__requiere_vencimiento=True
    )
    movimientos = MovimientoInventario.objects.all().order_by('-fecha')[:20]

    context = {
        'hoy': hoy,
        'piezas_stock_bajo': piezas_stock_bajo,
        'lotes_vencidos': lotes_vencidos,
        'lotes_por_vencer': lotes_por_vencer,
        'movimientos': movimientos,
    }

    template = get_template("reportes/reporte_pdf.html")
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Error generando PDF", status=500)
    return response


